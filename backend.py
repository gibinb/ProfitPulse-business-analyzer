import bcrypt
import jwt
import datetime
import io
import os
import smtplib
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from database import get_connection

SECRET_KEY = os.getenv("JWT_SECRET", "supersecretkey")
ADMIN_CODE  = os.getenv("ADMIN_CODE",  "ADMIN2024")


# ─────────────────────────── AUTH ────────────────────────────

def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt())

def check_password(password, hashed):
    if isinstance(hashed, str):
        hashed = hashed.encode()
    return bcrypt.checkpw(password.encode(), hashed)

def create_token(username):
    payload = {
        "username": username,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=8),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")

def verify_token(token):
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=["HS256"])["username"]
    except Exception:
        return None


# ─────────────────────────── USER ────────────────────────────

def register_user(username, gmail, password, admin_code=""):
    if not username or not gmail or not password:
        return False
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM users WHERE username=?", (username,))
    if cursor.fetchone():
        conn.close()
        return False
    role = "Admin" if admin_code == ADMIN_CODE else "Owner"
    cursor.execute(
        "INSERT INTO users (username, gmail, password, role) VALUES (?,?,?,?)",
        (username, gmail, hash_password(password), role),
    )
    conn.commit(); conn.close()
    return True

def get_user(username):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username, password FROM users WHERE username=?", (username,))
    user = cursor.fetchone()
    conn.close()
    return user

def get_profile(username):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username, gmail, role FROM users WHERE username=?", (username,))
    user = cursor.fetchone()
    conn.close()
    return user

def get_user_role(username):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE username=?", (username,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else "Owner"

def update_user_role(username, new_role):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET role=? WHERE username=?", (new_role, username))
    conn.commit(); conn.close()

def get_all_users():
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username, gmail, role FROM users ORDER BY username")
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_all_businesses():
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, business_name, owner_username, created_at FROM business ORDER BY created_at DESC"
    )
    rows = cursor.fetchall()
    conn.close()
    return rows


# ─────────────────────────── BUSINESS ────────────────────────

def create_business(owner, business_name):
    if not business_name:
        return
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO business (owner_username, business_name) VALUES (?,?)",
        (owner, business_name),
    )
    conn.commit(); conn.close()

def get_user_businesses(username):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, business_name FROM business WHERE owner_username=?", (username,)
    )
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_user_business(username):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM business WHERE owner_username=? LIMIT 1", (username,)
    )
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None

def get_accessible_businesses(username, role):
    """Owner sees their own businesses. Accountant/Staff see granted businesses."""
    conn   = get_connection()
    cursor = conn.cursor()
    if role == "Owner":
        cursor.execute(
            "SELECT id, business_name FROM business WHERE owner_username=?",
            (username,),
        )
    else:
        cursor.execute(
            """SELECT b.id, b.business_name
               FROM business b
               JOIN business_access ba ON b.id = ba.business_id
               WHERE ba.username=?""",
            (username,),
        )
    rows = cursor.fetchall()
    conn.close()
    return rows

def grant_business_access(username, business_id, granted_by):
    """Give a user access to a specific business."""
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM business_access WHERE username=? AND business_id=?",
        (username, business_id),
    )
    if not cursor.fetchone():
        cursor.execute(
            "INSERT INTO business_access (username, business_id, granted_by) VALUES (?,?,?)",
            (username, business_id, granted_by),
        )
        conn.commit()
    conn.close()

def revoke_business_access(username, business_id):
    """Remove a user's access to a specific business."""
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "DELETE FROM business_access WHERE username=? AND business_id=?",
        (username, business_id),
    )
    conn.commit(); conn.close()

def get_user_access_list(username):
    """Get list of business IDs a user currently has access to."""
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT business_id FROM business_access WHERE username=?",
        (username,),
    )
    rows = cursor.fetchall()
    conn.close()
    return [r[0] for r in rows]

def get_team_members(business_id):
    """Get all Accountant/Staff who have access to a specific business."""
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """SELECT u.username, u.gmail, u.role, ba.granted_at
           FROM users u
           JOIN business_access ba ON u.username = ba.username
           WHERE ba.business_id=?
           ORDER BY ba.granted_at DESC""",
        (business_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return rows

def owner_create_team_member(username, gmail, password, role, business_id, owner_username):
    """Owner creates Accountant or Staff and grants business access in one step."""
    if not register_user(username, gmail, password):
        return False, "Username already exists or invalid input."
    update_user_role(username, role)
    grant_business_access(username, business_id, owner_username)
    return True, "Team member created successfully!"


# ─────────────────────────── TRANSACTIONS ────────────────────

def save_sales(username, business_id, product, qty, amount, date, notes=""):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT unit_cost FROM inventory WHERE product=? AND business_id=? LIMIT 1",
        (product, business_id),
    )
    result     = cursor.fetchone()
    unit_cost  = result[0] if result else 0
    total_cogs = unit_cost * qty

    cursor.execute(
        """INSERT INTO transactions
           (username, type, amount, business_id, cogs, product, quantity, txn_date, notes)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (username, "Sales", amount, business_id, total_cogs,
         product, qty, str(date), notes),
    )
    cursor.execute(
        "UPDATE inventory SET quantity = quantity - ? WHERE product=? AND business_id=?",
        (qty, product, business_id),
    )
    cursor.execute(
        "INSERT INTO inventory_movements (business_id, product, change_qty, movement_type) VALUES (?,?,?,?)",
        (business_id, product, -qty, "OUT"),
    )
    conn.commit(); conn.close()

def save_expense(username, business_id, amount, category, date, notes=""):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """INSERT INTO transactions
           (username, type, amount, business_id, category, txn_date, notes)
           VALUES (?,?,?,?,?,?,?)""",
        (username, "Expense", amount, business_id, category, str(date), notes),
    )
    conn.commit(); conn.close()

def get_transactions(business_id, txn_type=None, limit=50):
    conn   = get_connection()
    cursor = conn.cursor()
    if txn_type:
        cursor.execute(
            """SELECT id, type, amount, category, product, quantity, txn_date, notes
               FROM transactions WHERE business_id=? AND type=?
               ORDER BY created_at DESC LIMIT ?""",
            (business_id, txn_type, limit),
        )
    else:
        cursor.execute(
            """SELECT id, type, amount, category, product, quantity, txn_date, notes
               FROM transactions WHERE business_id=?
               ORDER BY created_at DESC LIMIT ?""",
            (business_id, limit),
        )
    rows = cursor.fetchall()
    conn.close()
    return rows

def delete_transaction(transaction_id, business_id):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "DELETE FROM transactions WHERE id=? AND business_id=?",
        (transaction_id, business_id),
    )
    conn.commit(); conn.close()

def update_transaction(transaction_id, business_id, amount, notes):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE transactions SET amount=?, notes=? WHERE id=? AND business_id=?",
        (amount, notes, transaction_id, business_id),
    )
    conn.commit(); conn.close()

def get_expense_by_category(business_id):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """SELECT category, SUM(amount) FROM transactions
           WHERE business_id=? AND type='Expense' GROUP BY category""",
        (business_id,),
    )
    data = cursor.fetchall()
    conn.close()
    return data


# ─────────────────────────── PROFIT ──────────────────────────

def _period_filter(period):
    if period == "today":
        return " AND DATE(txn_date) = DATE('now')"
    if period == "week":
        return " AND strftime('%Y-%W', txn_date) = strftime('%Y-%W', 'now')"
    if period == "month":
        return " AND strftime('%Y-%m', txn_date) = strftime('%Y-%m', 'now')"
    return ""

def calculate_profit(business_id, period="all"):
    conn   = get_connection()
    cursor = conn.cursor()
    pf     = _period_filter(period)

    cursor.execute(
        f"SELECT SUM(amount) FROM transactions WHERE business_id=? AND type='Sales'{pf}",
        (business_id,),
    )
    sales = float(cursor.fetchone()[0] or 0)

    cursor.execute(
        f"SELECT SUM(amount) FROM transactions WHERE business_id=? AND type='Expense'{pf}",
        (business_id,),
    )
    expense = float(cursor.fetchone()[0] or 0)

    cursor.execute(
        f"SELECT SUM(cogs) FROM transactions WHERE business_id=? AND type='Sales'{pf}",
        (business_id,),
    )
    cogs = float(cursor.fetchone()[0] or 0)

    conn.close()
    return sales, expense, cogs, sales - expense - cogs

def get_sales_trend(business_id, period="month"):
    conn   = get_connection()
    cursor = conn.cursor()
    pf     = _period_filter(period)
    cursor.execute(
        f"""SELECT DATE(txn_date), SUM(amount), SUM(cogs)
            FROM transactions
            WHERE business_id=? AND type='Sales'{pf}
            GROUP BY DATE(txn_date) ORDER BY DATE(txn_date)""",
        (business_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return rows


# ─────────────────────────── INVENTORY ───────────────────────

def add_inventory(username, business_id, product, qty, cost, date, threshold):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """INSERT INTO inventory
           (username, product, quantity, unit_cost, purchase_date,
            business_id, low_stock_threshold)
           VALUES (?,?,?,?,?,?,?)""",
        (username, product, qty, cost, str(date), business_id, threshold),
    )
    cursor.execute(
        "INSERT INTO inventory_movements (business_id, product, change_qty, movement_type) VALUES (?,?,?,?)",
        (business_id, product, qty, "IN"),
    )
    conn.commit(); conn.close()

def get_inventory(business_id):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """SELECT product, quantity, unit_cost, low_stock_threshold, purchase_date
           FROM inventory WHERE business_id=? ORDER BY product""",
        (business_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_low_stock(business_id):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT product, quantity FROM inventory WHERE business_id=? AND quantity <= low_stock_threshold",
        (business_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_inventory_movements(business_id, limit=50):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """SELECT product, change_qty, movement_type, movement_date
           FROM inventory_movements WHERE business_id=?
           ORDER BY movement_date DESC LIMIT ?""",
        (business_id, limit),
    )
    rows = cursor.fetchall()
    conn.close()
    return rows

def compute_cogs(business_id):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT SUM(cogs) FROM transactions WHERE business_id=? AND type='Sales'",
        (business_id,),
    )
    result = cursor.fetchone()
    conn.close()
    return float(result[0] if result and result[0] else 0)


# ─────────────────────────── CSV ─────────────────────────────

def process_csv_profit(df):
    required = {"Date", "Product", "Quantity", "Selling_Price", "Cost_Price"}
    if not required.issubset(df.columns):
        return None, f"CSV must contain: {', '.join(required)}"
    df["Date"]    = pd.to_datetime(df["Date"])
    df["Revenue"] = df["Quantity"] * df["Selling_Price"]
    df["COGS"]    = df["Quantity"] * df["Cost_Price"]
    df["Profit"]  = df["Revenue"] - df["COGS"]
    daily = df.groupby("Date")[["Revenue", "COGS", "Profit"]].sum().reset_index()
    return (df, daily), None


# ─────────────────────────── AI INSIGHTS ─────────────────────

def generate_ai_insights(business_id, business_name):
    sales, expense, cogs, profit         = calculate_profit(business_id, "all")
    sales_m, expense_m, cogs_m, profit_m = calculate_profit(business_id, "month")
    low_stock  = get_low_stock(business_id)
    cat_data   = get_expense_by_category(business_id)
    trend_rows = get_sales_trend(business_id, "month")

    insights        = []
    recommendations = []

    if sales == 0:
        insights.append("No sales data found yet. Start adding transactions to see insights.")
        return insights, recommendations

    margin = (profit / sales) * 100
    if margin >= 30:
        insights.append(f"Strong profit margin of {margin:.1f}% - your business is highly profitable.")
    elif margin >= 10:
        insights.append(f"Moderate profit margin of {margin:.1f}% - there is room for improvement.")
        recommendations.append("Consider reducing operating expenses or increasing product prices slightly.")
    else:
        insights.append(f"Low profit margin of {margin:.1f}% - costs are eating into profits.")
        recommendations.append("Urgently review your largest expense categories and find areas to cut costs.")

    if expense > 0:
        expense_ratio = (expense / sales) * 100
        if expense_ratio > 60:
            insights.append(f"Expenses are {expense_ratio:.1f}% of sales - dangerously high.")
            recommendations.append("Your expenses are too high. Review Salary and Rent costs first.")
        elif expense_ratio > 40:
            insights.append(f"Expenses are {expense_ratio:.1f}% of sales - watch this closely.")
            recommendations.append("Keep expenses below 40% of sales for a healthier business.")
        else:
            insights.append(f"Expenses are well controlled at {expense_ratio:.1f}% of sales.")

    if cogs > 0:
        cogs_ratio = (cogs / sales) * 100
        if cogs_ratio > 60:
            insights.append(f"COGS is {cogs_ratio:.1f}% of revenue - very high cost of goods.")
            recommendations.append("Negotiate better prices with suppliers to reduce COGS.")
        elif cogs_ratio > 40:
            insights.append(f"COGS is {cogs_ratio:.1f}% of revenue - moderate.")
        else:
            insights.append(f"COGS is healthy at {cogs_ratio:.1f}% of revenue.")

    if cat_data:
        top_cat = max(cat_data, key=lambda x: x[1])
        insights.append(f"Highest expense category: {top_cat[0]} - Rs {float(top_cat[1]):,.0f}")
        if top_cat[0] in ("Salary", "Rent"):
            recommendations.append(f"{top_cat[0]} is your biggest cost. Ensure it is proportional to your revenue.")

    if trend_rows and len(trend_rows) >= 3:
        revenues = [float(r[1]) for r in trend_rows]
        avg_rev  = sum(revenues) / len(revenues)
        last_rev = revenues[-1]
        if last_rev > avg_rev * 1.1:
            insights.append(f"Sales are trending UP - last day Rs {last_rev:,.0f} vs avg Rs {avg_rev:,.0f}.")
        elif last_rev < avg_rev * 0.9:
            insights.append(f"Sales are trending DOWN - last day Rs {last_rev:,.0f} vs avg Rs {avg_rev:,.0f}.")
            recommendations.append("Sales are declining. Consider running a promotion or marketing campaign.")
        else:
            insights.append(f"Sales are stable - averaging Rs {avg_rev:,.0f} per day this month.")

    if low_stock:
        items = ", ".join([i[0] for i in low_stock])
        insights.append(f"{len(low_stock)} item(s) low on stock: {items}")
        recommendations.append(f"Restock {items} soon to avoid losing sales.")

    if profit_m > 0:
        insights.append(f"This month is profitable - net profit of Rs {profit_m:,.0f}.")
    elif profit_m < 0:
        insights.append(f"This month is loss-making - net loss of Rs {abs(profit_m):,.0f}.")
        recommendations.append("You are losing money this month. Immediate cost review recommended.")
    else:
        insights.append("Breaking even this month.")

    return insights, recommendations


# ─────────────────────────── EMAIL REPORT ────────────────────

def send_report_email(to_email, business_name, business_id, period, report_type="pdf"):
    SMTP_EMAIL    = os.getenv("SMTP_EMAIL",    "")
    SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")

    if not SMTP_EMAIL or not SMTP_PASSWORD:
        return "SMTP credentials not configured. Set SMTP_EMAIL and SMTP_PASSWORD environment variables."

    try:
        if report_type == "pdf":
            file_bytes = generate_pdf_report(business_name, business_id, period)
            filename   = f"{business_name}_report.pdf"
        else:
            file_bytes = generate_excel_report(business_name, business_id, period)
            filename   = f"{business_name}_report.xlsx"

        msg            = MIMEMultipart()
        msg["From"]    = SMTP_EMAIL
        msg["To"]      = to_email
        # ── no em dash, no emoji in subject ──
        msg["Subject"] = f"ProfitPulse: {business_name} - {period.replace('all', 'All Time').capitalize()} Business Report"

        body = (
            f"Hi,\n\n"
            f"Please find attached your {period.replace('all', 'All Time')} "
            f"business report for {business_name}.\n\n"
            f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} by ProfitPulse.\n\n"
            f"Best regards,\n"
            f"ProfitPulse\n"
        )
        msg.attach(MIMEText(body, "plain"))

        part = MIMEBase("application", "octet-stream")
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.sendmail(SMTP_EMAIL, to_email, msg.as_string())

        return True

    except smtplib.SMTPAuthenticationError:
        return "Gmail authentication failed. Use a Gmail App Password, not your regular password."
    except smtplib.SMTPException as e:
        return f"SMTP error: {str(e)}"
    except Exception as e:
        return f"Error: {str(e)}"


# ─────────────────────────── REPORT LOGGING ──────────────────

def log_report(business_id, report_type, file_url="local"):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO reports (business_id, report_type, file_url) VALUES (?,?,?)",
        (business_id, report_type, file_url),
    )
    conn.commit(); conn.close()

def get_report_logs(business_id=None):
    conn   = get_connection()
    cursor = conn.cursor()
    if business_id:
        cursor.execute(
            """SELECT r.id, b.business_name, r.report_type, r.file_url, r.generated_at
               FROM reports r JOIN business b ON r.business_id = b.id
               WHERE r.business_id=? ORDER BY r.generated_at DESC""",
            (business_id,),
        )
    else:
        cursor.execute(
            """SELECT r.id, b.business_name, r.report_type, r.file_url, r.generated_at
               FROM reports r JOIN business b ON r.business_id = b.id
               ORDER BY r.generated_at DESC"""
        )
    rows = cursor.fetchall()
    conn.close()
    return rows


# ─────────────────────────── SYSTEM SETTINGS ─────────────────

def get_system_settings():
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM system_settings")
    rows = cursor.fetchall()
    conn.close()
    return {row[0]: row[1] for row in rows}

def update_system_setting(key, value):
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO system_settings (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=?",
        (key, value, value),
    )
    conn.commit(); conn.close()


# ─────────────────────────── REPORTS ─────────────────────────

def generate_pdf_report(business_name, business_id, period):
    from fpdf import FPDF

    sales, expense, cogs, profit = calculate_profit(business_id, period)
    margin = (profit / sales * 100) if sales > 0 else 0

    # ── sanitize business_name — remove any non-Latin-1 characters ──
    def safe(text):
        return text.encode("latin-1", errors="replace").decode("latin-1")

    class PDF(FPDF):
        def header(self):
            self.set_fill_color(41, 128, 185)
            self.set_text_color(255, 255, 255)
            self.set_font("Arial", "B", 16)
            # ── em dash replaced with plain hyphen ──
            self.cell(0, 14, "ProfitPulse - Business Report",
                      border=0, ln=1, align="C", fill=True)
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "I", 8)
            self.set_text_color(128)
            self.cell(0, 10,
                      f"Page {self.page_no()} | Auto-generated by ProfitPulse",
                      align="C")

    pdf = PDF()
    pdf.add_page()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(40, 8, "Business:", ln=0)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 8, safe(business_name), ln=1)

    pdf.set_font("Arial", "B", 11)
    pdf.cell(40, 8, "Period:", ln=0)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 8, period.replace("all", "All Time").capitalize(), ln=1)

    pdf.set_font("Arial", "B", 11)
    pdf.cell(40, 8, "Generated:", ln=0)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 8, datetime.datetime.now().strftime("%Y-%m-%d  %H:%M"), ln=1)
    pdf.ln(4)

    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 10, "Financial Summary", ln=1)

    col_w = [110, 80]
    pdf.set_fill_color(41, 128, 185)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(col_w[0], 9, "Metric",       border=1, align="C", fill=True)
    pdf.cell(col_w[1], 9, "Amount (INR)", border=1, align="C", fill=True, ln=1)

    report_rows = [
        ("Total Sales",               f"Rs {sales:>12,.2f}"),
        ("Total Expenses",            f"Rs {expense:>12,.2f}"),
        ("Cost of Goods Sold (COGS)", f"Rs {cogs:>12,.2f}"),
        ("Net Profit",                f"Rs {profit:>12,.2f}"),
        ("Profit Margin",             f"{margin:.2f} %"),
    ]

    pdf.set_text_color(0, 0, 0)
    for i, (label, value) in enumerate(report_rows):
        is_profit = label == "Net Profit"
        pdf.set_fill_color(245, 245, 245) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        if is_profit:
            pdf.set_font("Arial", "B", 11)
            pdf.set_text_color(39, 174, 96) if profit >= 0 else pdf.set_text_color(231, 76, 60)
        else:
            pdf.set_font("Arial", "", 11)
        fill = i % 2 == 0
        pdf.cell(col_w[0], 9, label,  border=1, fill=fill)
        pdf.cell(col_w[1], 9, value,  border=1, align="R", fill=fill, ln=1)
        pdf.set_text_color(0, 0, 0)

    log_report(business_id, "PDF")
    return bytes(pdf.output())


def generate_excel_report(business_name, business_id, period):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    BLUE   = PatternFill("solid", fgColor="2980B9")
    WHITE  = Font(color="FFFFFF", bold=True)
    CENTER = Alignment(horizontal="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    sales, expense, cogs, profit = calculate_profit(business_id, period)
    margin = (profit / sales * 100) if sales > 0 else 0

    ws.merge_cells("A1:B1")
    # ── em dash replaced with plain hyphen ──
    ws["A1"] = f"ProfitPulse Report - {business_name}"
    ws["A1"].font      = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    ws.append(["Period",    period.replace("all", "All Time").capitalize()])
    ws.append(["Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])

    ws.append(["Metric", "Amount (INR)"])
    for cell in ws[ws.max_row]:
        cell.font = WHITE; cell.fill = BLUE; cell.alignment = CENTER

    for label, value in [
        ("Total Sales",     round(sales,   2)),
        ("Total Expenses",  round(expense, 2)),
        ("COGS",            round(cogs,    2)),
        ("Net Profit",      round(profit,  2)),
        ("Profit Margin %", round(margin,  2)),
    ]:
        ws.append([label, value])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    ws2     = wb.create_sheet("Transactions")
    headers = ["ID", "Type", "Amount", "Category", "Product", "Qty", "Date", "Notes"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = WHITE; cell.fill = BLUE

    conn   = get_connection()
    cursor = conn.cursor()
    pf     = _period_filter(period)
    cursor.execute(
        f"""SELECT id, type, amount, category, product, quantity, txn_date, notes
            FROM transactions WHERE business_id=?{pf}
            ORDER BY txn_date DESC""",
        (business_id,),
    )
    for row in cursor.fetchall():
        ws2.append(list(row))
    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    ws3         = wb.create_sheet("Inventory")
    inv_headers = ["Product", "Quantity", "Unit Cost", "Low Stock Threshold", "Purchase Date"]
    ws3.append(inv_headers)
    for cell in ws3[1]:
        cell.font = WHITE; cell.fill = BLUE

    cursor.execute(
        "SELECT product, quantity, unit_cost, low_stock_threshold, purchase_date FROM inventory WHERE business_id=?",
        (business_id,),
    )
    for row in cursor.fetchall():
        ws3.append(list(row))
    for i in range(1, len(inv_headers) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 20

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    log_report(business_id, "Excel")
    return buf.getvalue()